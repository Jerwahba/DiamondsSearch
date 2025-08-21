import json
import re
import sys
from groq import Groq
import pandas as pd
import requests
import schedule
import threading
import time
from io import BytesIO

# Global dataframe to hold the diamond data
diamond_df = None
DB_URL = "https://inventory-download.s3-accelerate.amazonaws.com/MID_White_INV.XLSX"

def download_database(url):
    """Downloads the excel file and loads it into a pandas DataFrame."""
    global diamond_df
    try:
        print("Downloading database...", file=sys.stderr)
        response = requests.get(url)
        response.raise_for_status()  # Raise an exception for bad status codes
        # Use BytesIO to read the excel file content from memory
        excel_file = BytesIO(response.content)
        
        # Read all sheets from the excel file with hyperlinks
        excel_file.seek(0)  # Reset file pointer
        all_sheets = pd.read_excel(excel_file, engine='openpyxl', sheet_name=None)
        
        # Also read the workbook to extract hyperlinks
        excel_file.seek(0)
        from openpyxl import load_workbook
        workbook = load_workbook(excel_file, data_only=True)
        
        # Combine all sheets into one DataFrame
        all_dataframes = []
        total_diamonds = 0
        
        for sheet_name, df in all_sheets.items():
            if not df.empty:
                print(f"Processing sheet: {sheet_name} with {len(df)} rows", file=sys.stderr)
                try:
                    # Pre-process column names: lowercase and replace spaces with underscores
                    df.columns = [col.strip().lower().replace(' ', '_').replace('.', '').replace('%', '_percent') for col in df.columns]
                    
                    # Extract hyperlinks for DDP column if it exists
                    ddp_columns = [col for col in df.columns if 'ddp' in col.lower() or 'link' in col.lower() or 'url' in col.lower()]
                    if ddp_columns:
                        print(f"  📋 Found DDP-related columns: {ddp_columns}", file=sys.stderr)
                        for ddp_col in ddp_columns:
                            df = extract_hyperlinks_from_sheet(df, workbook[sheet_name], ddp_col)
                    elif 'ddp' in df.columns or 'stone_details_link' in df.columns:
                        ddp_col = 'ddp' if 'ddp' in df.columns else 'stone_details_link'
                        df = extract_hyperlinks_from_sheet(df, workbook[sheet_name], ddp_col)
                    
                    # Check if this sheet has diamond-related columns (basic validation)
                    diamond_columns = ['shape', 'color', 'clarity', 'weight', 'carat', 'lab', 'gia', 'igi', 'hrd']
                    has_diamond_data = any(col in df.columns for col in diamond_columns)
                    
                    if has_diamond_data:
                        all_dataframes.append(df)
                        total_diamonds += len(df)
                        print(f"  ✓ Sheet '{sheet_name}' added with {len(df)} diamonds", file=sys.stderr)
                    else:
                        print(f"  ⚠ Sheet '{sheet_name}' skipped - no diamond data detected", file=sys.stderr)
                        
                except Exception as e:
                    print(f"  ✗ Error processing sheet '{sheet_name}': {e}", file=sys.stderr)
                    continue
        
        if all_dataframes:
            # Combine all dataframes
            diamond_df = pd.concat(all_dataframes, ignore_index=True)
            print(f"Database downloaded and loaded successfully. Total diamonds: {total_diamonds} from {len(all_dataframes)} sheets.", file=sys.stderr)
        else:
            print("No valid data found in any sheet.", file=sys.stderr)
            diamond_df = pd.DataFrame()
            
    except requests.exceptions.RequestException as e:
        print(f"Error downloading the database: {e}", file=sys.stderr)
    except Exception as e:
        print(f"Error processing the database file: {e}", file=sys.stderr)
        print(f"Error details: {str(e)}", file=sys.stderr)

def extract_hyperlinks_from_sheet(df, worksheet, ddp_column):
    """Extract hyperlinks from the DDP column and replace text with actual URLs."""
    try:
        # Find the column index for DDP
        ddp_col_idx = None
        for idx, col in enumerate(df.columns):
            if col == ddp_column:
                ddp_col_idx = idx
                break
        
        if ddp_col_idx is None:
            print(f"  ⚠ DDP column '{ddp_column}' not found in sheet", file=sys.stderr)
            return df
        
        # Extract hyperlinks from the worksheet
        hyperlinks = {}
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row), start=2):
            cell = row[ddp_col_idx]
            if cell.hyperlink:
                hyperlinks[row_idx - 2] = cell.hyperlink.target  # -2 because we start from row 2 and pandas starts from 0
        
        # Replace text with actual URLs in the DataFrame
        for idx, url in hyperlinks.items():
            if idx < len(df):
                df.iloc[idx, ddp_col_idx] = url
        
        print(f"  ✓ Extracted {len(hyperlinks)} hyperlinks from DDP column", file=sys.stderr)
        
        # If no hyperlinks found, try alternative method
        if len(hyperlinks) == 0:
            df = extract_links_alternative_method(df, ddp_column)
        
        return df
        
    except Exception as e:
        print(f"  ⚠ Error extracting hyperlinks: {e}", file=sys.stderr)
        # Try alternative method as fallback
        return extract_links_alternative_method(df, ddp_column)

def extract_links_alternative_method(df, ddp_column):
    """Alternative method to extract links from DDP column using regex patterns."""
    try:
        import re
        
        # Common patterns for diamond URLs
        url_patterns = [
            r'https?://[^\s]+',  # Standard URLs
            r'www\.[^\s]+',      # URLs starting with www
            r'[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}/[^\s]*',  # Domain patterns
        ]
        
        links_found = 0
        for idx, value in enumerate(df[ddp_column]):
            if pd.notna(value) and isinstance(value, str):
                for pattern in url_patterns:
                    match = re.search(pattern, value)
                    if match:
                        df.iloc[idx, df.columns.get_loc(ddp_column)] = match.group(0)
                        links_found += 1
                        break
        
        if links_found > 0:
            print(f"  ✓ Found {links_found} links using alternative method", file=sys.stderr)
        
        return df
        
    except Exception as e:
        print(f"  ⚠ Error in alternative link extraction: {e}", file=sys.stderr)
        return df


def schedule_download():
    """Schedules the database download."""
    download_database(DB_URL)  # Initial download
    
    # Print database statistics
    if diamond_df is not None and not diamond_df.empty:
        print(f"Database loaded successfully!", file=sys.stderr)
        print(f"Total diamonds: {len(diamond_df)}", file=sys.stderr)
        print(f"Columns available: {list(diamond_df.columns)}", file=sys.stderr)
        print(f"Sample data shape: {diamond_df.shape}", file=sys.stderr)
    else:
        print("Warning: Database is empty or not loaded properly.", file=sys.stderr)
    
    schedule.every(5).minutes.do(download_database, DB_URL)
    while True:
        schedule.run_pending()
        time.sleep(1)

class IA_Monitoring:
    def __init__(self, api_key=None):
        """Initialize with Groq API client."""
        # Use environment variable or default key for testing
        if api_key is None:
            import os
            api_key = os.getenv('GROQ_API_KEY', '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')
        """Initialize with Groq API client."""
        self.groq_client = Groq(api_key=api_key)
        self.model = "llama-3.1-8b-instant"

    def set_model(self, new_model):
        """Change le modèle Groq utilisé."""
        self.model = new_model

    def detect_multiple_diamonds(self, body):
        """Detect if the message contains multiple diamond requests and split them."""
        prompt = f"""
        Tu es un assistant qui détecte si un message contient **plusieurs demandes de diamants distinctes**.

        Voici le message : "{body}"
        
        Analyse le message et détermine s'il contient plusieurs demandes de diamants séparées.
        
        Exemples de messages avec plusieurs diamants :
        - "Je cherche un diamant de 2 carats ET un diamant de 3 carats"
        - "Je veux un diamant rond de 1.5 carats et un diamant princesse de 2 carats"
        - "Cherchez-moi un diamant VS1 et aussi un diamant SI1"
        - "Je cherche 2 diamants : un de 2 carats et un de 3 carats"
        - "F round GIA ET G princess IGI" (demandes courtes multiples)
        - "VS1 2 carats et SI1 1.5 carats" (demandes concises multiples)
        
        Si tu détectes plusieurs demandes, divise le message en parties séparées.
        
        Réponds en JSON avec cette structure :
        {{
            "has_multiple": true/false,
            "requests": [
                "première demande de diamant",
                "deuxième demande de diamant",
                ...
            ]
        }}
        
        Si il n'y a qu'une seule demande, réponds :
        {{
            "has_multiple": false,
            "requests": ["{body}"]
        }}
        
        IMPORTANT : Ne divise que si les demandes sont vraiment distinctes et séparées.
        """
        
        try:
            response = self.groq_client.chat.completions.create(
                model=self.model,
                messages=[
                    {
                        "role": "system",
                        "content": "Assistant qui analyse les messages et détecte les demandes multiples de diamants. Réponds uniquement en JSON valide."
                    },
                    {
                        "role": "user",
                        "content": prompt
                    }
                ],
                temperature=0,
                max_completion_tokens=1024,
                response_format={"type": "json_object"}
            )
            
            # Parse the JSON response
            json_response = response.choices[0].message.content
            return json.loads(json_response)
                
        except Exception as e:
            print(f"Error detecting multiple diamonds: {e}", file=sys.stderr)
            # Fallback: assume single request
            return {"has_multiple": False, "requests": [body]}

    def verify_message(self, body):
        # First, check for common short diamond request patterns
        if self._is_short_diamond_request(body):
            return "TRUE"
            
        prompt = f"""
        Tu es un assistant qui détecte si un message est **une demande explicite** de recherche ou d'achat d'un diamant.

        Voici le message : "{body}"
        
        Répond uniquement par "TRUE" si le message **contient une demande explicite et claire** de rechercher ou acheter un diamant, par exemple:
        - "Je cherche un diamant de 3 carats"
        - "Je veux acheter un diamant bleu"
        - "F round GIA" (demande courte avec critères)
        - "VS1 2 carats" (demande concise)
        - "round D color" (demande brève)
        - "GIA princess" (demande minimale)
        - "2 carats VS2" (demande compacte)
        - "F color round" (demande courte)
        
        Répond "FALSE" si le message est une simple mention, une description d'un diamant possédé, une expression de satisfaction ou autre.
        Exemple: "Regarde le diamant que j'ai acheté pour ma femme" → FALSE
        Exemple: "Bonjour" → FALSE
        Exemple: "Comment allez-vous?" → FALSE
        
        IMPORTANT : Les demandes courtes avec des critères de diamant (couleur, forme, laboratoire, etc.) sont considérées comme des demandes valides, même si elles sont très concises.
        """
        try:
            response = self.groq_client.chat.completions.create(
                model=self.model,
                messages=[
                    {
                        "role": "system",
                        "content": "Assistant de classification simple qui répond uniquement par TRUE ou FALSE."
                    },
                    {
                        "role": "user",
                        "content": prompt
                    }
                ],
                temperature=0,
                max_completion_tokens=10
            )
            return response.choices[0].message.content.strip()
        except Exception as e:
            print(f"Error calling Groq API: {e}", file=sys.stderr)
            return "FALSE"
    
    def _is_short_diamond_request(self, body):
        """Check if the message is a short diamond request using pattern matching."""
        import re
        
        # Clean the message
        clean_body = body.strip().upper()
        
        # Common diamond criteria patterns
        color_pattern = r'\b([D-Z])\b'  # D, E, F, G, H, I, J, K, L, M
        shape_pattern = r'\b(ROUND|PRINCESS|CUSHION|OVAL|EMERALD|PEAR|MARQUISE|ASSCHER|RADIANT|HEART)\b'
        clarity_pattern = r'\b(IF|VVS1|VVS2|VS1|VS2|SI1|SI2)\b'
        lab_pattern = r'\b(GIA|IGI|HRD)\b'
        weight_pattern = r'\b(\d+(?:\.\d+)?)\s*(?:CARAT|CT|CARATS)\b'
        
        # Check if the message contains at least 2 diamond criteria
        criteria_found = 0
        
        if re.search(color_pattern, clean_body):
            criteria_found += 1
        if re.search(shape_pattern, clean_body):
            criteria_found += 1
        if re.search(clarity_pattern, clean_body):
            criteria_found += 1
        if re.search(lab_pattern, clean_body):
            criteria_found += 1
        if re.search(weight_pattern, clean_body):
            criteria_found += 1
            
        # If we find at least 2 criteria and the message is short (less than 50 chars), it's likely a diamond request
        return criteria_found >= 2 and len(body.strip()) <= 50

    def generate_filter(self, body):
        prompt = f"""
        You are an expert jeweler's assistant. Your task is to extract diamond search criteria from a user's message and return them as a valid JSON object.

        The user message is: "{body}"

        Please extract the following attributes if they are mentioned in the message:
        - shape: The shape of the diamond (e.g., round, princess, cushion, oval, emerald, pear, marquise, asscher, radiant, heart). If the user mentions "round brilliant", just extract "round".
        - weight: The weight of the diamond in carats.
        - color: The color of the diamond (e.g., D, E, F, G, H, I, J, K, L, M).
        - clarity: The clarity of the diamond (e.g., IF, VVS1, VVS2, VS1, VS2, SI1, SI2).
        - lab: The grading laboratory (e.g., GIA, IGI, HRD).
        - cut: The cut grade of the diamond (e.g., EX, VG, G).
        - polish: The polish grade of the diamond (e.g., EX, VG, G).
        - symmetry: The symmetry grade of the diamond (e.g., EX, VG, G).
        - depth: The depth percentage of the diamond.
        - table: The table percentage of the diamond.
        - flour: The fluorescence of the diamond (e.g., F, M, N, SL).
        - measurements: The measurements of the diamond (e.g., "11.95x12.02x7.3").
        - sell_percent: The sell percentage (e.g., -23.63).
        - ppc: The price per carat.
        - total_price: The total price.
        - location: The location of the diamond (e.g., Hong Kong, New York).
        - ddp: The diamond detail page URL or link to the diamond.

        IMPORTANT:
        - Only include the fields that are explicitly mentioned in the user's request.
        - For numeric fields (weight, depth, table, sell_percent, ppc, total_price), you can extract a single value or a range (e.g., "between 2 and 3 carats"). If a range is given, use suffixes "_min" and "_max" (e.g. "weight_min": 2, "weight_max": 3). If a single value is given, use the key without a suffix (e.g. "weight": 2).
        - For clarity, if the user asks for "or better", just return the specified clarity level (e.g., "VS2 or better" -> "clarity": "VS2"). The search logic will handle the "or better" part.
        - Return ONLY the JSON object, without any other text, explanations, or markdown. Your response should be parsable by a JSON parser.
        - If no search criteria can be extracted, return an empty JSON object {{}}.
        - Pay special attention to short, concise requests like "F round GIA" which should extract: {{"color": "F", "shape": "round", "lab": "GIA"}}
        - Handle abbreviated requests like "VS1 2 carats" which should extract: {{"clarity": "VS1", "weight": 2}}
        """

        response = self.groq_client.chat.completions.create(
            model=self.model,
            messages=[
                {
                    "role": "system",
                    "content": "You are an AI assistant that extracts information and returns it ONLY in valid JSON format."
                },
                {
                    "role": "user",
                    "content": prompt
                }
            ],
            temperature=0,
            max_completion_tokens=1024,
            response_format={"type": "json_object"}
        )
        
        # Get response content from Groq
        response_content = response.choices[0].message.content
        
        # If AI fails to extract filters, try pattern matching for short requests
        if not response_content or response_content.strip() == '{}':
            return self._extract_filters_pattern_matching(body)
            
        return response_content
    
    def _extract_filters_pattern_matching(self, body):
        """Extract filters using pattern matching for short requests."""
        import re
        import json
        
        filters = {}
        clean_body = body.strip().upper()
        
        # Extract color
        color_match = re.search(r'\b([D-Z])\b', clean_body)
        if color_match:
            filters['color'] = color_match.group(1)
        
        # Extract shape
        shape_match = re.search(r'\b(ROUND|PRINCESS|CUSHION|OVAL|EMERALD|PEAR|MARQUISE|ASSCHER|RADIANT|HEART)\b', clean_body)
        if shape_match:
            filters['shape'] = shape_match.group(1).lower()
        
        # Extract clarity
        clarity_match = re.search(r'\b(IF|VVS1|VVS2|VS1|VS2|SI1|SI2)\b', clean_body)
        if clarity_match:
            filters['clarity'] = clarity_match.group(1)
        
        # Extract lab
        lab_match = re.search(r'\b(GIA|IGI|HRD)\b', clean_body)
        if lab_match:
            filters['lab'] = lab_match.group(1)
        
        # Extract weight
        weight_match = re.search(r'\b(\d+(?:\.\d+)?)\s*(?:CARAT|CT|CARATS)?\b', clean_body)
        if weight_match:
            try:
                filters['weight'] = float(weight_match.group(1))
            except ValueError:
                pass
        
        return json.dumps(filters)

def search_diamonds(filters):
    """Searches for diamonds in the DataFrame based on the provided filters."""
    global diamond_df
    if diamond_df is None:
        print("Database not loaded yet.", file=sys.stderr)
        return []

    results_df = diamond_df.copy()
    if not isinstance(results_df, pd.DataFrame) or results_df.empty:
        print(f"Database is empty. Shape: {diamond_df.shape if diamond_df is not None else 'None'}", file=sys.stderr)
        return []
    
    print(f"Searching in database with {len(results_df)} diamonds. Available columns: {list(results_df.columns)}", file=sys.stderr)

    # Ensure text columns are strings to prevent linter errors before filtering
    for col in ['shape', 'color', 'clarity', 'lab', 'cut', 'polish', 'symmetry', 'flour', 'location', 'measurements', 'ddp']:
        if col in results_df.columns:
            results_df[col] = results_df[col].astype(str)
            
    mask = pd.Series(True, index=results_df.index)

    clarity_order = ['SI2', 'SI1', 'VS2', 'VS1', 'VVS2', 'VVS1', 'IF']
    
    # Simple text filters (case-insensitive)
    text_filters = ['color', 'lab', 'cut', 'polish', 'symmetry', 'flour', 'location', 'ddp']
    for key in text_filters:
        if key in filters and filters[key] is not None:
            # Ensure the filter value is a valid string (not None, not empty, not numeric types that shouldn't be strings)
            filter_value = filters[key]
            
            # Skip if the value is None, empty, or not a string-like value
            if filter_value is None or filter_value == "":
                continue
                
            # Convert to string safely
            filter_value = str(filter_value).strip()
            
            # Skip if the converted string is empty or "nan" or "None"
            if filter_value == "" or filter_value.lower() in ['nan', 'none', 'null']:
                continue
                
            mask &= results_df[key].str.contains(filter_value, case=False, na=False)
    
    if 'shape' in filters and filters['shape'] is not None:
        # Use startswith for shape to match 'round' from 'round brilliant'
        shape_value = str(filters['shape']).strip().lower()
        if shape_value and shape_value not in ['nan', 'none', 'null']:
            mask &= results_df['shape'].str.lower().str.startswith(shape_value, na=False)
        
    if 'clarity' in filters and filters['clarity'] is not None:
        clarity_value = str(filters['clarity']).strip()
        if clarity_value and clarity_value.lower() not in ['nan', 'none', 'null']:
            try:
                clarity_index = clarity_order.index(clarity_value.upper())
                better_clarities = clarity_order[clarity_index:]
                mask &= results_df['clarity'].isin(better_clarities)
            except (ValueError, AttributeError):
                # If the clarity is not in our list, do a regular contains search
                mask &= results_df['clarity'].str.contains(clarity_value, case=False, na=False)

    # Numeric range filters
    numeric_filters = ['weight', 'depth', 'table', 'sell_percent', 'ppc', 'total_price']
    for key in numeric_filters:
        col_name = key
        if f"{key}_min" in filters:
            mask &= pd.to_numeric(results_df[col_name], errors='coerce') >= filters[f"{key}_min"]
        if f"{key}_max" in filters:
            mask &= pd.to_numeric(results_df[col_name], errors='coerce') <= filters[f"{key}_max"]
        if key in filters and f"{key}_min" not in filters and f"{key}_max" not in filters:
             mask &= pd.to_numeric(results_df[col_name], errors='coerce') == filters[key]

    # Handle measurements separately as it's a string match
    if 'measurements' in filters and filters['measurements'] is not None:
        measurements_value = str(filters['measurements']).strip()
        if measurements_value and measurements_value.lower() not in ['nan', 'none', 'null']:
            mask &= results_df['measurements'].str.contains(measurements_value, case=False, na=False)

    final_results_df = results_df[mask]

    results_to_return = final_results_df.head(10)
    # Replace pandas NaN with None for valid JSON
    results_with_none = pd.DataFrame(results_to_return.astype(object).where(pd.notnull(results_to_return), None))
    # After the where operation, we need to restore the column names
    results_with_none.columns = results_to_return.columns


    return results_with_none.to_dict('records')

def main():
    if len(sys.argv) < 2 or not sys.argv[1]:
        print(json.dumps({}))
        sys.exit(0)

    body = sys.argv[1]

    monitor = IA_Monitoring()

    # First, detect if there are multiple diamond requests
    multiple_detection = monitor.detect_multiple_diamonds(body)
    
    all_results = []
    all_filters = []
    
    # Process each request separately
    for i, request in enumerate(multiple_detection["requests"]):
        print(f"Processing request {i+1}: {request}", file=sys.stderr)
        
        verification_response = monitor.verify_message(request)
        cleaned_verification = re.sub(r'<think>.*?</think>', '', verification_response, flags=re.DOTALL).strip()

        if cleaned_verification != "TRUE":
            continue

        filter_response = monitor.generate_filter(request)
        if not filter_response:
            continue

        cleaned_filters_str = re.sub(r'<think>.*?</think>', '', filter_response, flags=re.DOTALL).strip()
        match = re.search(r'\{.*\}', cleaned_filters_str, re.DOTALL)
        if not match:
            continue

        json_str = match.group(0)

        try:
            filters = json.loads(json_str)
            if filters:
                if diamond_df is None:
                     print(json.dumps({"error": "Database not loaded yet. Please try again in a moment."}))
                     return
                search_results = search_diamonds(filters)
                
                # Add request context to filters and results
                filters_with_context = {
                    "request_text": request,
                    "request_number": i + 1,
                    **filters
                }
                
                # Add request context to each result
                results_with_context = []
                for result in search_results:
                    result_with_context = {
                        **result,
                        "request_number": i + 1,
                        "request_text": request
                    }
                    results_with_context.append(result_with_context)
                
                all_filters.append(filters_with_context)
                all_results.extend(results_with_context)
            else:
                continue
        except json.JSONDecodeError:
            continue

    # Prepare final output
    if all_filters and all_results:
        if multiple_detection["has_multiple"]:
            # Multiple requests - return array of filters
            output = {
                "multiple_requests": True,
                "total_requests": len(multiple_detection["requests"]),
                "filters": all_filters,
                "results": all_results
            }
        else:
            # Single request - return single filter object for compatibility
            output = {
                "multiple_requests": False,
                "total_requests": 1,
                "filters": all_filters[0],  # Return the single filter object
                "results": all_results
            }
        print(json.dumps(output, indent=4))
    else:
        print(json.dumps({}))


if __name__ == "__main__":
    # Start the background thread for downloading the database
    download_thread = threading.Thread(target=schedule_download, daemon=True)
    download_thread.start()

    # Give the initial download a moment to complete before processing requests
    time.sleep(10) # Adjust this as needed based on download speed

    main()
